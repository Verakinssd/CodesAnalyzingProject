import copy
import json
import openpyxl
import concurrent.futures
from openai import OpenAI
from sentence_transformers import SentenceTransformer
import numpy as np
import networkx as nx
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import community as commuinity_louvain
from collections import defaultdict


def read_submission_info_from_excel(filePath):
    workbook = openpyxl.load_workbook(filePath)
    sheet = workbook.active
    submission_list = []

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        Submission = {
            "学号": row[0],
            "姓名": row[1],
            "提交编号": row[2],
            "测评结果": row[3],
            "测试点详细": row[4],
            "得分": row[5],
            "语言": row[6],
            "代码长度": row[7],
            "运行时间": row[8],
            "运行内存": row[9],
            "提交时间": row[10],
            "代码": row[11].replace("_x000D_", "")
        }
        submission_list.append(Submission)

    return submission_list


def analyze_community(apiKey , String) :
    client = OpenAI(
        api_key=apiKey,
        base_url="https://api.deepseek.com",
    )
    
    response = client.chat.completions.create(
        model="deepseek-chat",
        messages=[
            {"role": "system", "content": "你是一个OJ评测结果分析总结系统"},
            {"role": "user", "content": String},
        ],
    )
    return response.choices[0].message.content


def analyze_code(apiKey, Question, sampleInputOutput, studentCode, dataRange, judgeAnswer, submission_id):
    client = OpenAI(
        api_key=apiKey,
        base_url="https://api.deepseek.com",
    )

    template = r"""
    有如下编程题
    {q}
    
    错误代码
    {c}

    样例数据
    {d}
    
    数据范围
    {range}
    
    测评结果
    {judge_answer}
    
    EXAMPLE JSON OUTPUT:
    {{
        "错误类型": "",
        "错误原因": ""
    }}
    """

    prompt = template.format(q=Question, c=studentCode, d=sampleInputOutput, range=dataRange, judge_answer=judgeAnswer)

    try:
        response = client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": "你是一个OJ评测机"},
                {"role": "user", "content": prompt},
            ],
            response_format={'type': 'json_object'}
        )

        if response and response.choices and response.choices[0].message.content:
            content = response.choices[0].message.content
            #print(f"AI 返回内容: {content}")
            try:
                data = json.loads(content)
            except json.JSONDecodeError as e:
                print(f"JSON 解析错误: {e}")
                data = {"错误类型": "JSON解析错误", "错误原因": str(e)}
        else:
            data = {"错误类型": "无响应", "错误原因": "API 未返回有效数据"}
    except Exception as e:
        print(f"API 调用失败: {e}")
        data = {"错误类型": "API调用失败", "错误原因": str(e)}
    
    # 添加提交编号
    data["submission_id"] = submission_id
    return data


def analyze_similarity(error_reasons):
    model = SentenceTransformer('lier007/xiaobu-embedding-v2')
    embeddings = model.encode(error_reasons, normalize_embeddings=True)
    similarity_matrix = np.dot(embeddings, embeddings.T)
    return similarity_matrix


def get_str(idx, analyzes) :
    return f"{analyzes[idx]}"
    
    
    
if __name__ == '__main__':
    excel_file = "E:\\Codes_error_analyzing_project\\C2\\test.xlsx"
    #excel path
    
    question = r"""
    Kanna 和朋友们一同经营着一家名为“星光咖啡馆”的小店。到月末的时候，她们需要统计整理本月的营业额。现在已知咖啡馆在上个月共售出了 n 杯咖啡，每杯咖啡的售价都是 p 元。请问上个月售出的这些咖啡为店里带来的营业额是多少。

    输入格式  
    一行，两个正整数 n, p，以空格分隔，含义如题目描述。

    输出格式  
    一行，一个整数，为题目所要求的营业额值。
    """

    sample_input_output = r"""
    输入数据：
    50 6

    输出数据：
    300
    """

    data_range = """
    保证 a, b在 int 范围内且为正整数
    """

    submissions = read_submission_info_from_excel(excel_file)
    api_key = "sk-d67797d1790a403082c4a34f387d3946"
    total_submissions = len(submissions)

    error_reasons = []  # 用于保存所有错误原因
    analyzes = [] # 用于保存所有分析结果

    with open("answer.jsonl", "w", encoding="utf-8") as f:
        futures = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            for i, submission in enumerate(submissions):
                print(f"正在处理第 {i + 1}/{total_submissions} 个提交...")
                t = copy.deepcopy(submission['代码'])
                t2 = copy.deepcopy(submission['测评结果'])
                futures.append(executor.submit(analyze_code, api_key, question, sample_input_output, t, data_range, t2, i))

            for (i, submission), future in zip(enumerate(submissions), futures):
                progress = (i + 1) / total_submissions * 100
                print(f"进度: {progress:.2f}%")

                try:
                    analysis = future.result()
                    analyzes.append(analysis)
                    #print(f"AI 分析结果：\n{analysis}")
                    error_reasons.append(analysis.get("错误原因", "未知错误"))

                    # 保存结果到文件
                    f.write(f"{json.dumps(analysis, ensure_ascii=False)}\n")
                    f.flush()
                except Exception as e:
                    print(f"分析提交 {i + 1} 时出错: {e}")
                    f.write(f"分析提交 {i + 1} 时出错: {e}\n")

    print("所有分析结果已保存到 answer.json 文件中")

    # 进行错误原因相似度分析
    if error_reasons:
        print("\n正在进行错误原因相似度分析...")
        similarity_matrix = analyze_similarity(error_reasons)
        print("相似度矩阵:")
        print(similarity_matrix)

        # 保存相似度矩阵到文件
        with open("similarity_matrix.json", "w", encoding="utf-8") as sim_file:
            sim_file.write(json.dumps(similarity_matrix.tolist(), ensure_ascii=False))
        print("相似度分析结果已保存到 similarity_matrix.json 文件中")
        
    # with open("similarity_matrix.json", "r", encoding="utf-8") as sim_file:
    #     similarity_matrix = eval(sim_file.read())
    
    # with open("answer.jsonl", "r", encoding="utf-8") as sim_file:
    #     analyzes = [json.loads(_) for _ in sim_file.readlines()]
    
    print(f"size = {len(analyzes)}")
    
    print("创建一个邻接矩阵")
    adj_matrix = np.array(similarity_matrix)
    print(adj_matrix.shape)
    print("从邻接矩阵生成Graph对象")
    G = nx.from_numpy_array(adj_matrix)
    
    
    # print("可视化图")
    # nx.draw(G, with_labels=True)
    # plt.show()
    
    print("计算最佳分区")
    partition = commuinity_louvain.best_partition(G)
    
    num_communities = len(set(partition.values()))
    print(f"社区数量: {num_communities}")
    print(f"社区划分结果: {partition}")
    
    communities = defaultdict(list) 
    for k,v in partition.items():
        communities[v].append(k)
    communities = dict(sorted(communities.items()))
        
    target_str = dict()
    for k,v in communities.items():
        target_str[k] = '\n'.join([get_str(_,analyzes) for _ in v])
    
    for k in communities.keys() :
        with open(f"community{k}.txt","w",encoding="utf-8") as f :
            f.write(f"社区{k}:\n{target_str[k]}\n")
            analyze_answer = analyze_community(api_key, target_str[k])
            f.write(analyze_answer)
            # print(f"社区{k}:\n{target_str[k]}\n")
            # print(analyze_answer)
        
    # pos = nx.spring_layout(G)
    
    # cmap = cm.get_cmap('viridis' , max(partition.values()) + 1)
    # nx.draw_networkx_nodes(G, pos, partition.keys(), node_size=40, cmap=cmap, node_color=list(partition.values()))
    # nx.draw_networkx_edges(G, pos, alpha=0.5)
    # plt.show()
    